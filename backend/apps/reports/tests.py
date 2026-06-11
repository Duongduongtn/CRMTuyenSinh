"""Tests cho app reports (Sprint 3 Tuần 6).

Cover:
- Permission IsAdminOrAccountant chặn user thường, cho admin + accountant.
- Revenue endpoint group-by-date đúng số tiền + count.
- Conversion endpoint tính rate lead → enrollment → paid.
- Excel export trả về file XLSX (content-type + header attachment), parse được bằng openpyxl.
"""
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from apps.courses.models import Course, VehicleClass
from apps.leads.models import Lead, LeadSource, LeadStatus
from apps.orders.models import Enrollment
from apps.payments.models import Payment, PaymentMethod, PaymentStatus

User = get_user_model()


def _make_user(username, group=None, superuser=False):
    if superuser:
        u = User.objects.create_superuser(username=username, password="pw-strong-123")
    else:
        u = User.objects.create_user(username=username, password="pw-strong-123")
    if group:
        g, _ = Group.objects.get_or_create(name=group)
        u.groups.add(g)
    return u


def _make_course(slug="b-mt", price="8000000"):
    return Course.objects.create(
        slug=slug,
        title=f"Khoá {slug}",
        vehicle_class=VehicleClass.B_MT,
        tuition_fee=Decimal(price),
        deposit_amount=Decimal("1000000"),
    )


def _make_enrollment(course, code, phone="0903456789"):
    return Enrollment.objects.create(
        code=code,
        course=course,
        student_name="Nguyễn Văn Test",
        student_phone=phone,
        tuition_fee=Decimal("8000000"),
        deposit_amount=Decimal("1000000"),
    )


class ReportPermissionTests(TestCase):
    """Endpoint báo cáo chỉ cho admin + accountant + superuser."""

    def setUp(self):
        self.url = reverse("report-revenue")

    def test_anonymous_blocked(self):
        resp = self.client.get(self.url)
        self.assertIn(resp.status_code, (401, 403))

    def test_normal_staff_blocked(self):
        _make_user("staff1", group="sale")
        self.client.login(username="staff1", password="pw-strong-123")
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 403)

    def test_accountant_allowed(self):
        _make_user("acc1", group="accountant")
        self.client.login(username="acc1", password="pw-strong-123")
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 200)

    def test_superuser_allowed(self):
        _make_user("super1", superuser=True)
        self.client.login(username="super1", password="pw-strong-123")
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 200)


class RevenueReportTests(TestCase):
    """Revenue endpoint group-by-date."""

    def setUp(self):
        _make_user("acc1", group="accountant")
        self.client.login(username="acc1", password="pw-strong-123")
        self.course = _make_course()
        self.enrollment = _make_enrollment(self.course, code="ORD-REV1")

    def test_revenue_groups_by_date(self):
        now = timezone.now()
        # 2 payment confirmed hôm nay + 1 hôm qua
        for amt in (Decimal("1000000"), Decimal("2000000")):
            p = Payment.objects.create(
                enrollment=self.enrollment,
                amount=amt,
                method=PaymentMethod.MANUAL,
                status=PaymentStatus.CONFIRMED,
                confirmed_at=now,
            )
            # confirmed_at có thể bị recompute trong save signal; ép lại.
            Payment.objects.filter(pk=p.pk).update(confirmed_at=now)
        p2 = Payment.objects.create(
            enrollment=self.enrollment,
            amount=Decimal("500000"),
            method=PaymentMethod.MANUAL,
            status=PaymentStatus.CONFIRMED,
            confirmed_at=now - timedelta(days=1),
        )
        Payment.objects.filter(pk=p2.pk).update(confirmed_at=now - timedelta(days=1))

        resp = self.client.get(reverse("report-revenue"))
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertIn("rows", data)
        self.assertIn("summary", data)
        # 2 ngày khác nhau
        self.assertGreaterEqual(len(data["rows"]), 2)
        # Tổng = 3.500.000
        self.assertEqual(Decimal(data["summary"]["total_amount"]), Decimal("3500000"))
        self.assertEqual(data["summary"]["total_count"], 3)

    def test_revenue_excludes_pending(self):
        Payment.objects.create(
            enrollment=self.enrollment,
            amount=Decimal("9999999"),
            method=PaymentMethod.MANUAL,
            status=PaymentStatus.PENDING,
            confirmed_at=None,
        )
        resp = self.client.get(reverse("report-revenue"))
        self.assertEqual(Decimal(resp.json()["summary"]["total_amount"]), Decimal("0"))


class ConversionReportTests(TestCase):
    """Lead → Enrollment → Paid conversion %."""

    def setUp(self):
        _make_user("admin1", group="admin")
        self.client.login(username="admin1", password="pw-strong-123")
        self.course = _make_course()

    def test_conversion_rates(self):
        # 4 lead, 2 enrollment, 1 paid
        for i in range(4):
            Lead.objects.create(
                name=f"Khách {i}",
                phone=f"090345677{i}",
                source=LeadSource.WEBSITE,
                status=LeadStatus.NEW,
            )
        e1 = _make_enrollment(self.course, code="ORD-CV1", phone="0903456770")
        e2 = _make_enrollment(self.course, code="ORD-CV2", phone="0903456771")
        Payment.objects.create(
            enrollment=e1,
            amount=Decimal("1000000"),
            method=PaymentMethod.MANUAL,
            status=PaymentStatus.CONFIRMED,
            confirmed_at=timezone.now(),
        )

        resp = self.client.get(reverse("report-conversion"))
        self.assertEqual(resp.status_code, 200)
        d = resp.json()
        self.assertEqual(d["leads"], 4)
        self.assertEqual(d["enrollments"], 2)
        self.assertEqual(d["paid"], 1)
        self.assertEqual(d["rate_lead_to_enrollment_pct"], 50.0)
        self.assertEqual(d["rate_enrollment_to_paid_pct"], 50.0)
        self.assertEqual(d["rate_overall_pct"], 25.0)


class RevenueExportXlsxTests(TestCase):
    """Excel export hợp lệ, parse được + dữ liệu khớp."""

    def setUp(self):
        _make_user("acc1", group="accountant")
        self.client.login(username="acc1", password="pw-strong-123")
        course = _make_course()
        enrollment = _make_enrollment(course, code="ORD-XLS1")
        now = timezone.now()
        p = Payment.objects.create(
            enrollment=enrollment,
            amount=Decimal("1234567"),
            method=PaymentMethod.MANUAL,
            status=PaymentStatus.CONFIRMED,
            confirmed_at=now,
        )
        Payment.objects.filter(pk=p.pk).update(confirmed_at=now)

    def test_export_returns_xlsx(self):
        from openpyxl import load_workbook

        resp = self.client.get(reverse("report-revenue-export"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", resp["Content-Disposition"])
        self.assertIn(".xlsx", resp["Content-Disposition"])
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.title, "Doanh thu")
        self.assertEqual(ws["A1"].value, "BÁO CÁO DOANH THU")
        # Header row 4: Ngày | Doanh thu xác nhận (VND) | Số giao dịch
        self.assertEqual(ws.cell(row=4, column=2).value, "Doanh thu xác nhận (VND)")
        # Có ít nhất 1 row dữ liệu (row 5) với amount = 1234567
        self.assertEqual(ws.cell(row=5, column=2).value, 1234567)
