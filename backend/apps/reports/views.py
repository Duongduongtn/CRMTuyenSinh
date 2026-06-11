"""View báo cáo doanh thu + conversion + Excel export.

Endpoint:
- GET /api/admin/reports/revenue?from=YYYY-MM-DD&to=YYYY-MM-DD
- GET /api/admin/reports/conversion?from=YYYY-MM-DD&to=YYYY-MM-DD
- GET /api/admin/reports/revenue/export.xlsx?from=&to=

Tham số `from` / `to` tùy chọn; mặc định 30 ngày gần nhất, timezone Asia/Ho_Chi_Minh.
Dữ liệu hiển thị format VN ở lớp render (Excel + summary string), JSON giữ số raw.
"""
from datetime import datetime, time, timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import Count, DecimalField, Sum, Value
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.throttling import ScopedRateThrottle
from rest_framework.views import APIView

from apps.leads.models import Lead
from apps.orders.models import Enrollment
from apps.payments.models import Payment, PaymentStatus

from .permissions import IsAdminOrAccountant


DEFAULT_RANGE_DAYS = 30
MAX_RANGE_DAYS = 366  # 1 năm cộng dôi 1 ngày; chống DoS gen file XLSX khổng lồ.


def _parse_range(request):
    """Trả `(from_dt, to_dt)` aware datetime. To bao gồm cuối ngày `to`.

    Validate: sai format → 400. Range > MAX_RANGE_DAYS → 400.
    """
    tz = timezone.get_current_timezone()
    today = timezone.localdate()
    default_from = today - timedelta(days=DEFAULT_RANGE_DAYS - 1)

    def _parse_date(s, fallback, field):
        if not s:
            return fallback
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValidationError({field: "Định dạng YYYY-MM-DD."}) from exc

    from_date = _parse_date(request.query_params.get("from"), default_from, "from")
    to_date = _parse_date(request.query_params.get("to"), today, "to")
    if to_date < from_date:
        from_date, to_date = to_date, from_date
    if (to_date - from_date).days > MAX_RANGE_DAYS:
        raise ValidationError(
            {"range": f"Khoảng tối đa {MAX_RANGE_DAYS} ngày."}
        )
    from_dt = timezone.make_aware(datetime.combine(from_date, time.min), tz)
    to_dt = timezone.make_aware(datetime.combine(to_date, time.max), tz)
    return from_date, to_date, from_dt, to_dt


def _revenue_rows(from_dt, to_dt):
    """List dict {date, confirmed_amount, confirmed_count} theo ngày."""
    qs = (
        Payment.objects
        .filter(status=PaymentStatus.CONFIRMED, confirmed_at__range=(from_dt, to_dt))
        .annotate(day=TruncDate("confirmed_at"))
        .values("day")
        .annotate(
            confirmed_amount=Coalesce(Sum("amount"), Value(Decimal("0"), output_field=DecimalField())),
            confirmed_count=Count("id"),
        )
        .order_by("day")
    )
    return list(qs)


class RevenueReportView(APIView):
    permission_classes = [IsAdminOrAccountant]

    def get(self, request, *args, **kwargs):
        from_date, to_date, from_dt, to_dt = _parse_range(request)
        rows = _revenue_rows(from_dt, to_dt)
        total_amount = sum((r["confirmed_amount"] for r in rows), Decimal("0"))
        total_count = sum(r["confirmed_count"] for r in rows)
        return Response({
            "from": from_date.isoformat(),
            "to": to_date.isoformat(),
            "rows": [
                {
                    "date": r["day"].isoformat(),
                    "confirmed_amount": str(r["confirmed_amount"]),
                    "confirmed_count": r["confirmed_count"],
                }
                for r in rows
            ],
            "summary": {
                "total_amount": str(total_amount),
                "total_count": total_count,
            },
        })


class ConversionReportView(APIView):
    permission_classes = [IsAdminOrAccountant]

    def get(self, request, *args, **kwargs):
        from_date, to_date, from_dt, to_dt = _parse_range(request)
        leads_n = Lead.objects.filter(created_at__range=(from_dt, to_dt)).count()
        enrollments_n = Enrollment.objects.filter(created_at__range=(from_dt, to_dt)).count()
        # Đếm enrollment TẠO TRONG range mà ĐÃ CÓ ít nhất 1 payment confirmed
        # (bất kể payment confirmed sau range). Mục đích: "% đơn cohort tháng này
        # đã đóng cọc", không phải "thu trong tháng" (cái đó dùng RevenueReport).
        paid_n = (
            Enrollment.objects
            .filter(
                created_at__range=(from_dt, to_dt),
                payments__status=PaymentStatus.CONFIRMED,
            )
            .distinct()
            .count()
        )

        def _rate(num, den):
            return round((num / den) * 100, 2) if den else 0.0

        return Response({
            "from": from_date.isoformat(),
            "to": to_date.isoformat(),
            "leads": leads_n,
            "enrollments": enrollments_n,
            "paid": paid_n,
            "rate_lead_to_enrollment_pct": _rate(enrollments_n, leads_n),
            "rate_enrollment_to_paid_pct": _rate(paid_n, enrollments_n),
            "rate_overall_pct": _rate(paid_n, leads_n),
        })


class RevenueExportXlsxView(APIView):
    """Excel export, format VN: ngày dd/mm/yyyy, số `1.234.567`.

    Throttle 10/giờ/user (scope `report_export`) chống DoS gen XLSX lớn — kết
    hợp với clamp `MAX_RANGE_DAYS` trong `_parse_range`.
    """

    permission_classes = [IsAdminOrAccountant]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "report_export"

    def get(self, request, *args, **kwargs):
        # Import openpyxl muộn để app load nhanh hơn (không cần openpyxl khi chỉ chạy API JSON).
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        from_date, to_date, from_dt, to_dt = _parse_range(request)
        rows = _revenue_rows(from_dt, to_dt)

        wb = Workbook()
        ws = wb.active
        ws.title = "Doanh thu"

        # Tiêu đề
        ws["A1"] = "BÁO CÁO DOANH THU"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Khoảng: {from_date.strftime('%d/%m/%Y')} - {to_date.strftime('%d/%m/%Y')}"
        ws["A2"].font = Font(italic=True)

        # Header
        headers = ["Ngày", "Doanh thu xác nhận (VND)", "Số giao dịch"]
        header_fill = PatternFill("solid", fgColor="E5F4EC")
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        total = Decimal("0")
        total_count = 0
        for i, r in enumerate(rows, start=5):
            ws.cell(row=i, column=1, value=r["day"].strftime("%d/%m/%Y"))
            amount_cell = ws.cell(row=i, column=2, value=float(r["confirmed_amount"]))
            amount_cell.number_format = "#,##0"
            ws.cell(row=i, column=3, value=r["confirmed_count"])
            total += r["confirmed_amount"]
            total_count += r["confirmed_count"]

        # Tổng
        total_row = 5 + len(rows) + 1
        ws.cell(row=total_row, column=1, value="Tổng").font = Font(bold=True)
        amount_total = ws.cell(row=total_row, column=2, value=float(total))
        amount_total.number_format = "#,##0"
        amount_total.font = Font(bold=True)
        count_total = ws.cell(row=total_row, column=3, value=total_count)
        count_total.font = Font(bold=True)

        # Width cột
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 16

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"doanh-thu_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
